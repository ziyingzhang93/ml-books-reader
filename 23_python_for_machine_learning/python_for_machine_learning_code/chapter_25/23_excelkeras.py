import random

import numpy as np
import openpyxl
from sklearn.datasets import fetch_openml
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense

# Read data from OpenML
dataset = fetch_openml("diabetes", version=1, as_frame=True, return_X_y=False)["frame"]
header = list(dataset.columns)
rows = dataset.to_numpy().tolist()

# Create Excel workbook and write data into the default worksheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Diabetes"
sheet.append(header)
for row in rows:
    sheet.append(row)
# Save
wb.save("MLM.xlsx")

# Create data generator for Keras classifier model
def datagen(batch_size):
    """A generator to produce samples from database
    """
    wb = openpyxl.load_workbook("MLM.xlsx", read_only=True)
    sheet = wb.active
    maxrow = sheet.max_row
    while True:
        # Read rows from Excel file
        X = []
        y = []
        for _ in range(batch_size):
            # data starts at row 2
            row_num = random.randint(2, maxrow)
            rowdata = [cell.value for cell in sheet[row_num]]
            X.append(rowdata[:-1])
            y.append(1 if rowdata[-1]=="tested_positive" else 0)
        yield np.asarray(X), np.asarray(y)

# create binary classification model
model = Sequential()
model.add(Dense(16, input_dim=8, activation='relu'))
model.add(Dense(8, activation='relu'))
model.add(Dense(1, activation='sigmoid'))
model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])

# train model
history = model.fit(datagen(32), epochs=5, steps_per_epoch=20)
