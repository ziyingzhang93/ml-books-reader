import pandas as pd
from sklearn.datasets import fetch_openml
import openpyxl

# Read dataset from OpenML
dataset = fetch_openml("diabetes", version=1, as_frame=True, return_X_y=False)["frame"]
header = list(dataset.columns)
data = dataset.to_numpy().tolist()

# Create Excel workbook and write data into the default worksheet
wb = openpyxl.Workbook()
sheet = wb.active # use the default worksheet
sheet.title = "Diabetes"
for n,colname in enumerate(header):
    sheet.cell(row=1, column=1+n, value=colname)
for n,row in enumerate(data):
    for m,cell in enumerate(row):
        sheet.cell(row=2+n, column=1+m, value=cell)
# Save
wb.save("MLM.xlsx")
